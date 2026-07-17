"""Excel styling for the per-storm exposure workbook attached to alert emails.

VENDORED COPY — the upstream source of truth is
ds-storm-impact-harmonisation/src/source_exposure/style.py (the styling of
the historical archive workbook). Keep the two in sync so the email
attachment and the archive read as the same product.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Book palette (book/custom.css --bs-primary + hover; cosmo body greys)
BRAND = "55B284"
BRAND_DARK = "3E8F6B"
BAND = "EAF4EE"        # very light green, alternating rows
INK = "333333"
MUTE = "888888"
WHITE = "FFFFFF"
HAIR = "D3D3D3"

_HEADER_FILL = PatternFill("solid", fgColor=BRAND)
_HEADER_FONT = Font(color=WHITE, bold=True, size=11)
_BAND_FILL = PatternFill("solid", fgColor=BAND)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HAIR_SIDE = Side(style="thin", color=HAIR)
_BORDER = Border(left=_HAIR_SIDE, right=_HAIR_SIDE,
                 top=_HAIR_SIDE, bottom=_HAIR_SIDE)

MONEY = "#,##0"        # population exposures
PLAIN = "0"            # ids / seasons (no thousands separator, no decimals)


def style_data_sheet(ws, money_cols=(), plain_cols=(), widths=None,
                     band=True, hidden=()):
    """Header styling, freeze, filter, number formats, widths, zebra rows.
    `hidden` = header names to collapse (data kept, column hidden)."""
    maxrow, maxcol = ws.max_row, ws.max_column
    headers = {ws.cell(1, c).value: c for c in range(1, maxcol + 1)}

    for c in range(1, maxcol + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
        cell.alignment, cell.border = _CENTER, _BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(maxcol)}{maxrow}"
    ws.sheet_view.showGridLines = False

    def fmt(names, code):
        for h in names:
            c = headers.get(h)
            if not c:
                continue
            for r in range(2, maxrow + 1):
                ws.cell(r, c).number_format = code

    fmt(money_cols, MONEY)
    fmt(plain_cols, PLAIN)

    if band:
        for r in range(2, maxrow + 1, 2):
            for c in range(1, maxcol + 1):
                ws.cell(r, c).fill = _BAND_FILL

    for c in range(1, maxcol + 1):
        letter = get_column_letter(c)
        h = str(ws.cell(1, c).value or "")
        ws.column_dimensions[letter].width = (
            widths.get(h) if widths and h in widths
            else min(max(len(h) + 3, 11), 24))
        if h in hidden:
            ws.column_dimensions[letter].hidden = True


def build_readme(ws, blocks):
    """Render a cover sheet. `blocks` = list of (kind, text); kind in
    {title, subtitle, meta, h2, body, bullet, gap}."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 112
    r = 1
    for kind, text in blocks:
        cell = ws.cell(r, 1, text)
        if kind == "title":
            cell.font = Font(color=BRAND_DARK, bold=True, size=22)
            ws.row_dimensions[r].height = 30
        elif kind == "subtitle":
            cell.font = Font(color=INK, italic=True, size=13)
            ws.row_dimensions[r].height = 20
        elif kind == "meta":
            cell.font = Font(color=MUTE, size=10)
        elif kind == "h2":
            cell.font = Font(color=WHITE, bold=True, size=12)
            cell.fill = PatternFill("solid", fgColor=BRAND)
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=1)
            ws.row_dimensions[r].height = 22
        elif kind == "bullet":
            cell.value = "•  " + text
            cell.font = Font(color=INK, size=11)
            cell.alignment = Alignment(wrap_text=True, vertical="top",
                                       indent=1)
            ws.row_dimensions[r].height = max(16, 15 * (len(text) // 95 + 1))
        elif kind == "gap":
            pass
        else:  # body
            cell.font = Font(color=INK, size=11)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = max(16, 15 * (len(text) // 95 + 1))
        r += 1
